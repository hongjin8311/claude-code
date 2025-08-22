#!/usr/bin/env python3
"""
Azure OpenAI Resource Scanner
Scans all Azure OpenAI resources and collects model deployment information with PTU usage metrics.
"""

import json
import csv
from datetime import datetime, timedelta
from typing import List, Dict, Any
import argparse
import sys

from azure.identity import DefaultAzureCredential, InteractiveBrowserCredential
from azure.mgmt.cognitiveservices import CognitiveServicesManagementClient
from azure.mgmt.resource import ResourceManagementClient, SubscriptionClient
from azure.mgmt.monitor import MonitorManagementClient
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment


class AzureOpenAIScanner:
    def __init__(self, tenant_id: str = None):
        """Initialize the scanner with Azure credentials."""
        if tenant_id:
            self.credential = InteractiveBrowserCredential(tenant_id=tenant_id)
        else:
            self.credential = DefaultAzureCredential()
        
        self.subscription_client = SubscriptionClient(self.credential)
        self.openai_resources = []
        
    def get_subscriptions(self) -> List[str]:
        """Get all available subscriptions."""
        subscriptions = []
        try:
            for sub in self.subscription_client.subscriptions.list():
                subscriptions.append(sub.subscription_id)
                print(f"Found subscription: {sub.display_name} ({sub.subscription_id})")
        except Exception as e:
            print(f"Error getting subscriptions: {e}")
        
        return subscriptions
    
    def get_ptu_metrics(self, subscription_id: str, resource_group: str, resource_name: str, deployment_name: str) -> Dict[str, float]:
        """Get PTU usage metrics for a specific deployment over the last 7 days."""
        try:
            monitor_client = MonitorManagementClient(self.credential, subscription_id)
            
            # Resource ID for the cognitive services account
            resource_id = f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}/providers/Microsoft.CognitiveServices/accounts/{resource_name}"
            
            # Time range for the last 7 days
            end_time = datetime.utcnow()
            start_time = end_time - timedelta(days=7)
            
            # Define metrics to retrieve
            metric_names = "ProvisionedManagedThroughputUtilization"
            
            try:
                metrics_data = monitor_client.metrics.list(
                    resource_uri=resource_id,
                    timespan=f"{start_time.isoformat()}/{end_time.isoformat()}",
                    interval='PT1H',  # 1 hour intervals
                    metricnames=metric_names,
                    aggregation='Average,Maximum',
                    filter=f"DeploymentName eq '{deployment_name}'"
                )
                
                avg_values = []
                max_values = []
                
                for metric in metrics_data.value:
                    if metric.name.value == "ProvisionedManagedThroughputUtilization":
                        for timeseries in metric.timeseries:
                            for data_point in timeseries.data:
                                if data_point.average is not None:
                                    avg_values.append(data_point.average)
                                if data_point.maximum is not None:
                                    max_values.append(data_point.maximum)
                
                return {
                    'ptu_avg_7days': sum(avg_values) / len(avg_values) if avg_values else 0.0,
                    'ptu_max_7days': max(max_values) if max_values else 0.0
                }
                
            except Exception as e:
                print(f"Error getting PTU metrics for {deployment_name}: {e}")
                return {'ptu_avg_7days': 0.0, 'ptu_max_7days': 0.0}
                
        except Exception as e:
            print(f"Error initializing monitor client: {e}")
            return {'ptu_avg_7days': 0.0, 'ptu_max_7days': 0.0}

    def get_openai_resources(self, subscription_id: str) -> List[Dict[str, Any]]:
        """Get all OpenAI resources in a subscription."""
        resources = []
        try:
            resource_client = ResourceManagementClient(self.credential, subscription_id)
            cognitive_client = CognitiveServicesManagementClient(self.credential, subscription_id)
            
            # Get all cognitive services accounts
            for account in cognitive_client.accounts.list():
                if account.kind.lower() == 'openai':
                    resource_info = {
                        'subscription_id': subscription_id,
                        'resource_group': account.id.split('/')[4],
                        'name': account.name,
                        'location': account.location,
                        'sku': account.sku.name if account.sku else None,
                        'endpoint': account.properties.endpoint if account.properties else None,
                        'deployments': []
                    }
                    
                    # Get deployments for this OpenAI resource
                    try:
                        deployments = cognitive_client.deployments.list(
                            resource_group_name=resource_info['resource_group'],
                            account_name=account.name
                        )
                        
                        for deployment in deployments:
                            deployment_info = {
                                'name': deployment.name,
                                'model': deployment.properties.model.name if deployment.properties and deployment.properties.model else None,
                                'model_version': deployment.properties.model.version if deployment.properties and deployment.properties.model else None,
                                'sku_name': deployment.sku.name if deployment.sku else None,
                                'sku_capacity': deployment.sku.capacity if deployment.sku else None,
                                'scale_type': deployment.properties.scale_settings.scale_type if deployment.properties and deployment.properties.scale_settings else None
                            }
                            
                            # Get PTU metrics for this deployment
                            print(f"Getting PTU metrics for {deployment.name}...")
                            ptu_metrics = self.get_ptu_metrics(
                                subscription_id, 
                                resource_info['resource_group'], 
                                account.name, 
                                deployment.name
                            )
                            deployment_info.update(ptu_metrics)
                            
                            resource_info['deployments'].append(deployment_info)
                            
                    except Exception as e:
                        print(f"Error getting deployments for {account.name}: {e}")
                    
                    resources.append(resource_info)
                    print(f"Found OpenAI resource: {account.name} in {account.location}")
                    
        except Exception as e:
            print(f"Error scanning subscription {subscription_id}: {e}")
            
        return resources
    
    def scan_all_resources(self) -> List[Dict[str, Any]]:
        """Scan all subscriptions for OpenAI resources."""
        all_resources = []
        subscriptions = self.get_subscriptions()
        
        if not subscriptions:
            print("No subscriptions found or accessible.")
            return all_resources
            
        for subscription_id in subscriptions:
            print(f"\nScanning subscription: {subscription_id}")
            resources = self.get_openai_resources(subscription_id)
            all_resources.extend(resources)
            
        return all_resources
    
    def save_to_json(self, resources: List[Dict[str, Any]], filename: str = None):
        """Save resources to JSON file."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"azure_openai_resources_{timestamp}.json"
            
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump({
                'scan_date': datetime.now().isoformat(),
                'total_resources': len(resources),
                'resources': resources
            }, f, indent=2, ensure_ascii=False)
            
        print(f"JSON report saved to: {filename}")
        return filename
    
    def save_to_excel(self, resources: List[Dict[str, Any]], filename: str = None):
        """Save resources to Excel file with separate sheets for each resource group."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"azure_openai_resources_{timestamp}.xlsx"
        
        # Prepare data for DataFrame
        all_data = []
        for resource in resources:
            if resource['deployments']:
                for deployment in resource['deployments']:
                    row = {
                        'Subscription ID': resource['subscription_id'],
                        'Resource Group': resource['resource_group'],
                        'Resource Name': resource['name'],
                        'Location': resource['location'],
                        'Resource SKU': resource['sku'],
                        'Endpoint': resource['endpoint'],
                        'Deployment Name': deployment['name'],
                        'Model Name': deployment['model'],
                        'Model Version': deployment['model_version'],
                        'Deployment SKU': deployment['sku_name'],
                        'SKU Capacity': deployment['sku_capacity'],
                        'Scale Type': deployment['scale_type'],
                        'PTU Avg (7 days)': round(deployment.get('ptu_avg_7days', 0.0), 2),
                        'PTU Max (7 days)': round(deployment.get('ptu_max_7days', 0.0), 2)
                    }
                    all_data.append(row)
            else:
                # Resource without deployments
                row = {
                    'Subscription ID': resource['subscription_id'],
                    'Resource Group': resource['resource_group'],
                    'Resource Name': resource['name'],
                    'Location': resource['location'],
                    'Resource SKU': resource['sku'],
                    'Endpoint': resource['endpoint'],
                    'Deployment Name': '',
                    'Model Name': '',
                    'Model Version': '',
                    'Deployment SKU': '',
                    'SKU Capacity': '',
                    'Scale Type': '',
                    'PTU Avg (7 days)': 0.0,
                    'PTU Max (7 days)': 0.0
                }
                all_data.append(row)
        
        if not all_data:
            print("No data to export to Excel.")
            return None
            
        # Create DataFrame
        df = pd.DataFrame(all_data)
        
        # Group by resource group
        resource_groups = df['Resource Group'].unique()
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_data = []
        
        for rg in sorted(resource_groups):
            rg_data = df[df['Resource Group'] == rg]
            total_resources = rg_data['Resource Name'].nunique()
            total_deployments = len(rg_data[rg_data['Deployment Name'] != ''])
            avg_ptu_usage = rg_data['PTU Avg (7 days)'].mean()
            max_ptu_usage = rg_data['PTU Max (7 days)'].max()
            
            summary_data.append({
                'Resource Group': rg,
                'Total Resources': total_resources,
                'Total Deployments': total_deployments,
                'Avg PTU Usage (7 days)': round(avg_ptu_usage, 2),
                'Max PTU Usage (7 days)': round(max_ptu_usage, 2)
            })
        
        summary_df = pd.DataFrame(summary_data)
        
        # Add summary data to summary sheet
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            summary_ws.append(r)
        
        # Format summary sheet
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="366092")
        for cell in summary_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths for summary
        for column in summary_ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            summary_ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Create sheets for each resource group
        for rg in sorted(resource_groups):
            # Clean sheet name (Excel sheet names have restrictions)
            sheet_name = rg.replace('/', '_').replace('\\', '_')[:31]  # Excel sheet name limit
            
            rg_data = df[df['Resource Group'] == rg]
            
            # Create new sheet
            ws = wb.create_sheet(sheet_name)
            
            # Add data to sheet
            for r in dataframe_to_rows(rg_data, index=False, header=True):
                ws.append(r)
            
            # Format header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filename)
        print(f"Excel report saved to: {filename}")
        return filename

    def save_to_csv(self, resources: List[Dict[str, Any]], filename: str = None):
        """Save resources to CSV file."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"azure_openai_resources_{timestamp}.csv"
            
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                'subscription_id', 'resource_group', 'resource_name', 'location', 'sku',
                'endpoint', 'deployment_name', 'model_name', 'model_version', 
                'sku_name', 'sku_capacity', 'scale_type', 'ptu_avg_7days', 'ptu_max_7days'
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for resource in resources:
                if resource['deployments']:
                    for deployment in resource['deployments']:
                        row = {
                            'subscription_id': resource['subscription_id'],
                            'resource_group': resource['resource_group'],
                            'resource_name': resource['name'],
                            'location': resource['location'],
                            'sku': resource['sku'],
                            'endpoint': resource['endpoint'],
                            'deployment_name': deployment['name'],
                            'model_name': deployment['model'],
                            'model_version': deployment['model_version'],
                            'sku_name': deployment['sku_name'],
                            'sku_capacity': deployment['sku_capacity'],
                            'scale_type': deployment['scale_type'],
                            'ptu_avg_7days': deployment.get('ptu_avg_7days', 0.0),
                            'ptu_max_7days': deployment.get('ptu_max_7days', 0.0)
                        }
                        writer.writerow(row)
                else:
                    # Resource without deployments
                    row = {
                        'subscription_id': resource['subscription_id'],
                        'resource_group': resource['resource_group'],
                        'resource_name': resource['name'],
                        'location': resource['location'],
                        'sku': resource['sku'],
                        'endpoint': resource['endpoint'],
                        'deployment_name': '',
                        'model_name': '',
                        'model_version': '',
                        'sku_name': '',
                        'sku_capacity': '',
                        'scale_type': '',
                        'ptu_avg_7days': 0.0,
                        'ptu_max_7days': 0.0
                    }
                    writer.writerow(row)
                    
        print(f"CSV report saved to: {filename}")
        return filename
    
    def print_summary(self, resources: List[Dict[str, Any]]):
        """Print a summary of found resources."""
        print(f"\n{'='*80}")
        print("AZURE OPENAI RESOURCES SUMMARY WITH PTU METRICS")
        print(f"{'='*80}")
        print(f"Total Resources Found: {len(resources)}")
        
        if not resources:
            print("No Azure OpenAI resources found.")
            return
            
        total_deployments = sum(len(r['deployments']) for r in resources)
        print(f"Total Deployments: {total_deployments}")
        
        # Calculate PTU metrics summary
        all_ptu_avg = []
        all_ptu_max = []
        for resource in resources:
            for deployment in resource['deployments']:
                ptu_avg = deployment.get('ptu_avg_7days', 0.0)
                ptu_max = deployment.get('ptu_max_7days', 0.0)
                if ptu_avg > 0:
                    all_ptu_avg.append(ptu_avg)
                if ptu_max > 0:
                    all_ptu_max.append(ptu_max)
        
        if all_ptu_avg:
            print(f"PTU Usage Summary (Last 7 Days):")
            print(f"  Average PTU Usage: {sum(all_ptu_avg)/len(all_ptu_avg):.2f}%")
            print(f"  Highest PTU Usage: {max(all_ptu_max):.2f}%")
            print(f"  Deployments with PTU data: {len(all_ptu_avg)}")
        
        # Group by location
        locations = {}
        for resource in resources:
            location = resource['location']
            if location not in locations:
                locations[location] = 0
            locations[location] += 1
            
        print(f"\nResources by Location:")
        for location, count in sorted(locations.items()):
            print(f"  {location}: {count}")
        
        # Group by resource group
        resource_groups = {}
        for resource in resources:
            rg = resource['resource_group']
            if rg not in resource_groups:
                resource_groups[rg] = 0
            resource_groups[rg] += 1
            
        print(f"\nResources by Resource Group:")
        for rg, count in sorted(resource_groups.items()):
            print(f"  {rg}: {count}")
            
        # Group by model
        models = {}
        model_ptu_data = {}
        for resource in resources:
            for deployment in resource['deployments']:
                model = deployment['model']
                if model:
                    if model not in models:
                        models[model] = 0
                        model_ptu_data[model] = {'avg': [], 'max': []}
                    models[model] += 1
                    
                    # Collect PTU data for this model
                    ptu_avg = deployment.get('ptu_avg_7days', 0.0)
                    ptu_max = deployment.get('ptu_max_7days', 0.0)
                    if ptu_avg > 0:
                        model_ptu_data[model]['avg'].append(ptu_avg)
                    if ptu_max > 0:
                        model_ptu_data[model]['max'].append(ptu_max)
                    
        if models:
            print(f"\nDeployments by Model (with PTU metrics):")
            for model, count in sorted(models.items()):
                avg_ptus = model_ptu_data[model]['avg']
                max_ptus = model_ptu_data[model]['max']
                
                avg_ptu_str = f"{sum(avg_ptus)/len(avg_ptus):.1f}%" if avg_ptus else "N/A"
                max_ptu_str = f"{max(max_ptus):.1f}%" if max_ptus else "N/A"
                
                print(f"  {model}: {count} deployments (Avg PTU: {avg_ptu_str}, Max PTU: {max_ptu_str})")


def main():
    parser = argparse.ArgumentParser(description='Scan Azure OpenAI resources with PTU metrics')
    parser.add_argument('--tenant-id', help='Azure tenant ID for authentication')
    parser.add_argument('--output-json', help='JSON output filename')
    parser.add_argument('--output-csv', help='CSV output filename')
    parser.add_argument('--output-excel', help='Excel output filename')
    parser.add_argument('--no-summary', action='store_true', help='Skip printing summary')
    parser.add_argument('--excel-only', action='store_true', help='Only generate Excel file (skip JSON and CSV)')
    
    args = parser.parse_args()
    
    try:
        print("Initializing Azure OpenAI Scanner with PTU metrics...")
        scanner = AzureOpenAIScanner(tenant_id=args.tenant_id)
        
        print("Scanning for Azure OpenAI resources and collecting PTU metrics...")
        resources = scanner.scan_all_resources()
        
        if not args.no_summary:
            scanner.print_summary(resources)
        
        # Save results
        saved_files = []
        
        if args.excel_only:
            # Only generate Excel file
            excel_file = scanner.save_to_excel(resources, args.output_excel)
            if excel_file:
                saved_files.append(f"Excel: {excel_file}")
        else:
            # Generate all formats
            json_file = scanner.save_to_json(resources, args.output_json)
            csv_file = scanner.save_to_csv(resources, args.output_csv)
            excel_file = scanner.save_to_excel(resources, args.output_excel)
            
            saved_files.append(f"JSON: {json_file}")
            saved_files.append(f"CSV: {csv_file}")
            if excel_file:
                saved_files.append(f"Excel: {excel_file}")
        
        print(f"\nScan completed successfully!")
        print(f"Results saved to:")
        for file_info in saved_files:
            print(f"  - {file_info}")
        
    except KeyboardInterrupt:
        print("\nScan interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"Error during scan: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()